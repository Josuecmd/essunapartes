import io
import os
from typing import List
from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import StreamingResponse, HTMLResponse
from pydantic import BaseModel, Field
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = FastAPI(
    title="Sistema de Parte General - Brigada de Guardiamarinas",
    description="Backend para el conteo, control nominal y exportación del parte militar.",
    version="1.0.0"
)

# --- INTERFAZ VISUAL (HTML/CSS/JS DIRECTO) ---
HTML_TABLERO = """
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Tablero de Control - Parte General de Guardiamarinas</title>
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/css/all.min.css" rel="stylesheet">
    <style>
        :root {
            --primary: #1e293b;
            --secondary: #0f172a;
            --accent: #2563eb;
            --success: #16a34a;
            --danger: #dc2626;
            --bg: #f8fafc;
        }
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background-color: var(--bg);
            margin: 0; padding: 0; color: #334155;
        }
        .header {
            background: linear-gradient(135deg, var(--primary), var(--secondary));
            color: white; padding: 20px; text-align: center;
            box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);
        }
        .header h1 { margin: 0; font-size: 24px; letter-spacing: 0.5px; }
        .header p { margin: 5px 0 0 0; opacity: 0.8; font-size: 13px; }
        .container {
            max-width: 1200px; margin: 30px auto; padding: 0 20px;
            display: grid; grid-template-columns: 1fr 2fr; gap: 30px;
        }
        @media (max-width: 900px) { .container { grid-template-columns: 1fr; } }
        .card {
            background: white; border-radius: 12px; padding: 25px;
            box-shadow: 0 4px 6px -1px rgba(0,0,0,0.05); border: 1px solid #e2e8f0;
            height: fit-content;
        }
        .stats-grid { display: grid; grid-template-columns: repeat(3, 1fr); gap: 15px; margin-bottom: 25px; }
        .stat-card {
            background: #f1f5f9; padding: 15px; border-radius: 8px;
            text-align: center; border-left: 4px solid var(--accent);
        }
        .stat-card.success { border-left-color: var(--success); }
        .stat-card.danger { border-left-color: var(--danger); }
        .stat-number { font-size: 22px; font-weight: bold; color: var(--secondary); }
        .stat-label { font-size: 11px; color: #64748b; font-weight: 600; text-transform: uppercase; }
        .form-group { margin-bottom: 15px; }
        label { display: block; font-weight: 600; margin-bottom: 6px; font-size: 14px; color: var(--primary); }
        input, select {
            width: 100%; padding: 11px; border: 1px solid #cbd5e1;
            border-radius: 6px; box-sizing: border-box; font-size: 14px; background: #fff;
        }
        .btn {
            background-color: var(--accent); color: white; border: none;
            padding: 13px; font-size: 14px; font-weight: 600; border-radius: 6px;
            cursor: pointer; width: 100%; transition: background 0.2s;
            display: flex; justify-content: center; align-items: center; gap: 8px;
        }
        .btn:hover { background-color: #1d4ed8; }
        .btn-success { background-color: var(--success); margin-top: 15px; }
        .btn-success:hover { background-color: #15803d; }
        table { width: 100%; border-collapse: collapse; margin-top: 15px; font-size: 14px; }
        th, td { padding: 12px 10px; text-align: left; border-bottom: 1px solid #e2e8f0; }
        th { background-color: #f1f5f9; color: var(--primary); font-weight: 600; }
        .status-badge { padding: 4px 10px; border-radius: 12px; font-size: 11px; font-weight: bold; display: inline-block; }
        .badge-fila { background: #dcfce7; color: #15803d; }
        .badge-comision { background: #dbeafe; color: #1e40af; }
        .badge-other { background: #fef3c7; color: #d97706; }
        .status-container { margin-top: 15px; display: none; padding: 12px; border-radius: 6px; font-size: 13px; }
        .status-loading { background: #eff6ff; color: #1d4ed8; }
    </style>
</head>
<body>
    <div class="header">
        <h1><i class="fa-solid fa-shield-halved"></i> TABLERO DE CONTROL DE PARTES</h1>
        <p>Brigada de Guardiamarinas | Sistema Integrado de Fuerza General</p>
    </div>
    <div class="container">
        <div class="card">
            <h3><i class="fa-solid fa-user-plus"></i> Registrar Novedad</h3>
            <div class="form-group">
                <label>Nombre y Apellido</label>
                <input type="text" id="alumnoNombre" placeholder="Ej: G/M Juan Pérez">
            </div>
            <div class="form-group">
                <label>Sección / Curso</label>
                <select id="alumnoSeccion">
                    <option value="Primer Año Alfa">Primer Año Alfa</option>
                    <option value="Primer Año Bravo">Primer Año Bravo</option>
                    <option value="Primer Año Charlie">Primer Año Charlie</option>
                    <option value="Primer Año Delta">Primer Año Delta</option>
                    <option value="Primer Año Echo">Primer Año Echo</option>
                    <option value="Segundo Año Alfa">Segundo Año Alfa</option>
                    <option value="Segundo Año Bravo">Segundo Año Bravo</option>
                    <option value="Tercer Año Alfa">Tercer Año Alfa</option>
                    <option value="Tercer Año Bravo">Tercer Año Bravo</option>
                    <option value="Tercer Año Abastecimientos">Tercer Año Abastecimientos</option>
                    <option value="Cuarto Año Alfa">Cuarto Año Alfa</option>
                    <option value="Cuarto Año Bravo">Cuarto Año Bravo</option>
                    <option value="Cuarto Año Abastecimientos">Cuarto Año Abastecimientos</option>
                </select>
            </div>
            <div class="form-group">
                <label>Estado de Fuerza</label>
                <select id="alumnoEstado">
                    <option value="FILA">FILA (Presente)</option>
                    <option value="COMISIÓN">COMISIÓN</option>
                    <option value="EXC. LITERA">EXC. LITERA</option>
                    <option value="EXC. FORMACIÓN">EXC. FORMACIÓN</option>
                    <option value="EXC. POR REVALIDAR">EXC. POR REVALIDAR</option>
                    <option value="DESCANSO DOMICILIO">DESCANSO DOMICILIO</option>
                </select>
            </div>
            <button class="btn" onclick="agregarAlumno()"><i class="fa-solid fa-plus"></i> Insertar a la Lista</button>
            <hr style="margin: 25px 0; border: 0; border-top: 1px solid #e2e8f0;">
            <div class="form-group">
                <label>Nombre del Documento</label>
                <input type="text" id="fileName" value="Parte_General_Brigada">
            </div>
            <button class="btn btn-success" onclick="exportarParte()"><i class="fa-solid fa-file-excel"></i> Generar Excel</button>
            <div id="statusBox" class="status-container status-loading"></div>
        </div>
        <div class="card">
            <div class="stats-grid">
                <div class="stat-card"><div class="stat-number" id="totalAlumnos">0</div><div class="stat-label">Alta General</div></div>
                <div class="stat-card success"><div class="stat-number" id="totalFila">0</div><div class="stat-label">En Fila</div></div>
                <div class="stat-card danger"><div class="stat-number" id="totalNovedades">0</div><div class="stat-label">Novedades</div></div>
            </div>
            <h3><i class="fa-solid fa-list-check"></i> Fuerza Nominal</h3>
            <div style="overflow-x: auto;">
                <table>
                    <thead><tr><th>Nombre</th><th>Sección</th><th>Estado</th><th style="text-align:center;">Acción</th></tr></thead>
                    <tbody id="tablaAlumnos">
                        <tr><td colspan="4" style="text-align:center; color:#94a3b8; padding:20px;">No hay personal registrado.</td></tr>
                    </tbody>
                </table>
            </div>
        </div>
    </div>
    <script>
        let listaGuardiamarinas = [];
        function actualizarContadores() {
            document.getElementById('totalAlumnos').innerText = listaGuardiamarinas.length;
            const fila = listaGuardiamarinas.filter(a => a.estado === 'FILA').length;
            document.getElementById('totalFila').innerText = fila;
            document.getElementById('totalNovedades').innerText = listaGuardiamarinas.length - fila;
        }
        function agregarAlumno() {
            const nombre = document.getElementById('alumnoNombre').value.trim();
            const seccion = document.getElementById('alumnoSeccion').value;
            const estado = document.getElementById('alumnoEstado').value;
            if (!nombre) return alert("Ingrese el nombre");
            listaGuardiamarinas.push({ id: listaGuardiamarinas.length + 1, nombre, seccion, estado });
            document.getElementById('alumnoNombre').value = '';
            renderTabla();
        }
        function eliminarAlumno(id) {
            listaGuardiamarinas = listaGuardiamarinas.filter(a => a.id !== id);
            renderTabla();
        }
        function renderTabla() {
            const tbody = document.getElementById('tablaAlumnos');
            if (listaGuardiamarinas.length === 0) {
                tbody.innerHTML = `<tr><td colspan="4" style="text-align:center; color:#94a3b8; padding:20px;">No hay personal registrado.</td></tr>`;
                actualizarContadores(); return;
            }
            tbody.innerHTML = '';
            listaGuardiamarinas.forEach(a => {
                let bc = 'badge-other';
                if (a.estado === 'FILA') bc = 'badge-fila';
                if (a.estado === 'COMISIÓN') bc = 'badge-comision';
                tbody.innerHTML += `<tr><td><strong>${a.nombre}</strong></td><td>${a.seccion}</td><td><span class="status-badge ${bc}">${a.estado}</span></td>
                <td style="text-align:center;"><button style="background:none;border:none;color:var(--danger);cursor:pointer;" onclick="eliminarAlumno(${a.id})"><i class="fa-solid fa-trash-can"></i></button></td></tr>`;
            });
            actualizarContadores();
        }
        async function exportarParte() {
            if (listaGuardiamarinas.length === 0) return alert("Lista vacía");
            const filename = document.getElementById('fileName').value.trim() || 'Parte_General_Brigada';
            const statusBox = document.getElementById('statusBox');
            statusBox.style.display = "block";
            statusBox.innerHTML = 'Generando archivo...';
            try {
                const response = await fetch('/api/parte/procesar-y-exportar', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ nombre_archivo: filename, brigada: listaGuardiamarinas })
                });
                if (!response.ok) throw new Error("Error de cuadre en servidor");
                const blob = await response.blob();
                const url = window.URL.createObjectURL(blob);
                const a = document.createElement('a'); a.href = url; a.download = filename + ".xlsx";
                document.body.appendChild(a); a.click(); a.remove();
                statusBox.style.display = "none";
            } catch (e) {
                statusBox.style.background = "#fee2e2"; statusBox.style.color = "#991b1b";
                statusBox.innerHTML = "Error al procesar: " + e.message;
            }
        }
    </script>
</body>
</html>
"""

@app.get("/", response_class=HTMLResponse)
async def leer_tablero():
    return HTMLResponse(content=HTML_TABLERO)

# --- CONFIGURACIÓN DE CONSTANTES MILITARES ---
SECCIONES_VALIDAS = [
    "Primer Año Alfa", "Primer Año Bravo", "Primer Año Charlie", "Primer Año Delta", "Primer Año Echo",
    "Segundo Año Alfa", "Segundo Año Bravo",
    "Tercer Año Alfa", "Tercer Año Bravo", "Tercer Año Abastecimientos",
    "Cuarto Año Alfa", "Cuarto Año Bravo", "Cuarto Año Abastecimientos"
]

ESTADOS_VALIDOS = ["FILA", "COMISIÓN", "EXC. LITERA", "EXC. FORMACIÓN", "EXC. POR REVALIDAR", "DESCANSO DOMICILIO"]

# --- MODELOS DE DATOS (PYDANTIC) ---
class Guardiamarina(BaseModel):
    id: int
    nombre: str = Field(..., example="Guardiamarina Juan Pérez")
    seccion: str = Field(..., example="Primer Año Alfa")
    estado: str = Field(default="FILA", example="FILA")

class ParteRequest(BaseModel):
    nombre_archivo: str = Field(default="Parte_General_Brigada", example="Parte_Diario_07_Junio")
    brigada: List[Guardiamarina]

# --- FUNCIONES DE PROCESAMIENTO Y FORMATEO ---
def generar_excel_profesional(df_nominal: pd.DataFrame, df_resumen: pd.DataFrame) -> io.BytesIO:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_resumen.to_excel(writer, sheet_name='Resumen Numérico', index=True)
        df_nominal.to_excel(writer, sheet_name='Listado Nominal', index=False)
        workbook = writer.book
        
        ws_resumen = workbook['Resumen Numérico']
        ws_resumen.views.sheetView[0].showGridLines = True
        
        font_header = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        font_body = Font(name='Arial', size=10)
        font_total = Font(name='Arial', size=10, bold=True)
        fill_header = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid') 
        fill_total_row = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid') 
        border_thin = Side(border_style="thin", color="D9D9D9")
        border_double = Side(border_style="double", color="000000")
        box_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
        total_border = Border(top=border_thin, bottom=border_double)

        ws_resumen['A1'] = "Secciones / Cursos"
        for col in range(1, ws_resumen.max_column + 1):
            cell = ws_resumen.cell(row=1, column=col)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for row in range(2, ws_resumen.max_row + 1):
            is_total_row = (ws_resumen.cell(row=row, column=1).value == "TOTAL GENERAL")
            for col in range(1, ws_resumen.max_column + 1):
                cell = ws_resumen.cell(row=row, column=col)
                cell.font = font_total if (is_total_row or col == ws_resumen.max_column) else font_body
                cell.border = total_border if is_total_row else box_border
                if is_total_row:
                    cell.fill = fill_total_row
                cell.alignment = Alignment(horizontal='center' if col > 1 else 'left')

        for col in ws_resumen.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_resumen.column_dimensions[col_letter].width = max(max_len + 3, 12)

        ws_nominal = workbook['Listado Nominal']
        ws_nominal.views.sheetView[0].showGridLines = True
        for col in range(1, ws_nominal.max_column + 1):
            cell = ws_nominal.cell(row=1, column=col)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal='left', vertical='center')
        for row in range(2, ws_nominal.max_row + 1):
            for col in range(1, ws_nominal.max_column + 1):
                cell = ws_nominal.cell(row=row, column=col)
                cell.font = font_body
                cell.border = box_border
        for col in ws_nominal.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_nominal.column_dimensions[col_letter].width = max(max_len + 4, 15)

    output.seek(0)
    return output

@app.post("/api/parte/procesar-y-exportar")
async def procesar_y_exportar_parte(request: ParteRequest):
    if not request.brigada:
        raise HTTPException(status_code=400, detail="La lista de la brigada está vacía.")
    datos_nominales = [g.model_dump() for g in request.brigada]
    df_nominal = pd.DataFrame(datos_nominales)

    secciones_invalidas = set(df_nominal['seccion']) - set(SECCIONES_VALIDAS)
    if secciones_invalidas:
        raise HTTPException(status_code=400, detail=f"Secciones inválidas: {list(secciones_invalidas)}")
    estados_invalidos = set(df_nominal['estado']) - set(ESTADOS_VALIDOS)
    if estados_invalidos:
        raise HTTPException(status_code=400, detail=f"Estados inválidos: {list(estados_invalidos)}")

    conteo_real = df_nominal.groupby(['seccion', 'estado']).size().unstack(fill_value=0)
    matriz_final = conteo_real.reindex(index=SECCIONES_VALIDAS, columns=ESTADOS_VALIDOS, fill_value=0)
    matriz_final['TOTAL ALTA'] = matriz_final.sum(axis=1)
    matriz_final.loc['TOTAL GENERAL'] = matriz_final.sum(axis=0)

    excel_binario = generar_excel_profesional(df_nominal, matriz_final)
    nombre_descarga = f"{request.nombre_archivo.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        excel_binario,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre_descarga}"}
    )

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)